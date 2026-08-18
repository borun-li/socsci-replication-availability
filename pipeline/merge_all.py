#!/usr/bin/env python3
"""Merge the 5 SocSci availability tables into one canonical v3.2 archive.
b1,b2,b4,b5 = canonical schema; b3 = 18-col batch-3 schema (mapped + DOI fetched from Crossref).
Output: socsci_all_v3.xlsx (sheet 'availability'), sorted by paper_id, + a 'batch' column.
"""
import openpyxl, re, urllib.request, urllib.parse, json, sys, time

BASE='artifacts/output/package_avai'
CANON=['doi','paper_id','title','authors','published_date','submission_date','article_url',
       'in_scope','qualitative','data','code','data_and_code','neither','data_gated',
       'data_source_apply_at','package_location','path_to_package','coverage_checked','notes']
B3MAP={'paper_id':'paper_id','title':'title','author(s)':'authors','published_date':'published_date',
       'submission_date':'submission_date','article_url':'article_url','in_scope (Y/NA)':'in_scope',
       'qualitative (Y/N)':'qualitative','data (Y/N)':'data','code (Y/N)':'code','data+code':'data_and_code',
       'neither':'neither','data_gated (Y/N)':'data_gated','data_source / apply_at':'data_source_apply_at',
       'package_location':'package_location','path_to_package':'path_to_package',
       'coverage_checked':'coverage_checked','notes':'notes'}

_doi_cache={}
def crossref_doi(title):
    if not title: return ''
    if title in _doi_cache: return _doi_cache[title]
    q=urllib.parse.urlencode({'query.bibliographic':title,'rows':5})
    req=urllib.request.Request(f"https://api.crossref.org/works?{q}",
        headers={'User-Agent':'socsci-availability/1.0 (mailto:borun.li@icloud.com)'})
    doi=''
    try:
        d=json.load(urllib.request.urlopen(req,timeout=25))
        for it in d.get('message',{}).get('items',[]):
            ct=' '.join(it.get('container-title',[]) or [])
            if 'sociological science' in ct.lower(): doi=it.get('DOI',''); break
    except Exception: doi=''
    _doi_cache[title]=doi
    return doi

def read_canon(f, batch):
    ws=openpyxl.load_workbook(f'{BASE}/{f}')['availability']
    idx={c.value:i for i,c in enumerate(ws[1])}
    out=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        if not r[idx['paper_id']]: continue
        row={k:r[idx[k]] for k in CANON}
        row['batch']=batch
        out.append(row)
    return out

def read_b3(f, batch):
    wb=openpyxl.load_workbook(f'{BASE}/{f}'); ws=wb['Borun']
    idx={c.value:i for i,c in enumerate(ws[1])}
    out=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        if not r[idx['paper_id']]: continue
        row={}
        for src,dst in B3MAP.items(): row[dst]=r[idx[src]]
        row['doi']=crossref_doi(row.get('title'))
        row['batch']=batch
        out.append(row)
    return out

def pid_key(p):
    m=re.match(r'SS0*(\d+)', p or ''); return int(m.group(1)) if m else 0

def main():
    rows=[]
    rows+=read_canon('socsci_batch1_v3.xlsx','1')
    rows+=read_canon('socsci_batch2_v3.xlsx','2')
    rows+=read_b3('Borun_batch_3_result.xlsx','3')
    rows+=read_canon('socsci_batch4.xlsx','4')
    rows+=read_canon('socsci_batch5.xlsx','5')
    rows+=read_canon('socsci_batch6_gap.xlsx','6')
    rows.sort(key=lambda r: pid_key(r['paper_id']))
    cols=CANON+['batch']
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='availability'
    ws.append(cols)
    for r in rows: ws.append([r.get(k) for k in cols])
    ws.freeze_panes='A2'
    out=f'{BASE}/socsci_all_v3.xlsx'
    wb.save(out)
    # report
    b3doi=sum(1 for r in rows if r['batch']=='3' and r.get('doi'))
    print(f"merged {len(rows)} rows -> {out}")
    print(f"batch counts:", {b:sum(1 for r in rows if r['batch']==b) for b in '12345'})
    print(f"batch-3 DOIs resolved: {b3doi}/80")
    y=[r for r in rows if r['in_scope']=='Y']
    av=[r for r in y if str(r['data'])=='Y' or str(r['code'])=='Y']
    print(f"TOTAL in_scope Y={len(y)} | availability={len(av)}/{len(y)}={len(av)/len(y)*100:.1f}%")

if __name__=='__main__': main()
